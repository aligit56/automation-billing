import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def generate_excel(filepath="dummy_attendance_records.xlsx"):
    # Define the data
    data = [
        # Clean Records (Standard Pass)
        {
            "period_key": "2026-08", "employee_code": "EMP-1001", "full_name": "Alice Smith", 
            "email": "alice@axian.com", "location_code": "US-MAIN", "days_worked": 22.0, 
            "approved_leaves": 0.0, "unapproved_absences": 0.0, "status": "ACTIVE", 
            "verification_status": "PENDING", "employee_notes": ""
        },
        # Approved Leave Case
        {
            "period_key": "2026-08", "employee_code": "EMP-1002", "full_name": "Bob Jones", 
            "email": "bob@axian.com", "location_code": "US-MAIN", "days_worked": 20.0, 
            "approved_leaves": 2.0, "unapproved_absences": 0.0, "status": "ACTIVE", 
            "verification_status": "PENDING", "employee_notes": "Sick leave approved"
        },
        # Working Days Mismatch Exception
        {
            "period_key": "2026-08", "employee_code": "EMP-1004", "full_name": "Diana Prince", 
            "email": "diana@axian.com", "location_code": "US-MAIN", "days_worked": 18.0, 
            "approved_leaves": 0.0, "unapproved_absences": 0.0, "status": "ACTIVE", 
            "verification_status": "PENDING", "employee_notes": "Missed some days"
        },
        # Unapproved Absence Violation
        {
            "period_key": "2026-08", "employee_code": "EMP-1003", "full_name": "Charlie Brown", 
            "email": "charlie@axian.com", "location_code": "US-MAIN", "days_worked": 19.0, 
            "approved_leaves": 0.0, "unapproved_absences": 3.0, "status": "ACTIVE", 
            "verification_status": "PENDING", "employee_notes": "No call no show"
        },
        # Inactive Employee Entry
        {
            "period_key": "2026-08", "employee_code": "EMP-1005", "full_name": "Eve Adams", 
            "email": "eve@axian.com", "location_code": "US-MAIN", "days_worked": 22.0, 
            "approved_leaves": 0.0, "unapproved_absences": 0.0, "status": "INACTIVE", 
            "verification_status": "PENDING", "employee_notes": "Was terminated mid-month"
        },
        # Unknown Headcount Entry (Ghost Employee)
        {
            "period_key": "2026-08", "employee_code": "EMP-9999", "full_name": "Ghost Employee", 
            "email": "ghost@axian.com", "location_code": "US-MAIN", "days_worked": 22.0, 
            "approved_leaves": 0.0, "unapproved_absences": 0.0, "status": "ACTIVE", 
            "verification_status": "PENDING", "employee_notes": "Not in master DB"
        },
    ]

    df = pd.DataFrame(data)

    # Write to Excel
    writer = pd.ExcelWriter(filepath, engine='openpyxl')
    df.to_excel(writer, sheet_name="Attendance_Summary", index=False)
    writer.close()

    # Load workbook for formatting
    wb = load_workbook(filepath)
    ws = wb["Attendance_Summary"]

    # Formatting headers
    header_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid") # Dark Blue
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Number formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = '0.0'

    # Auto-fit columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) if column_cells else 10
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Data Validation
    dv = DataValidation(type="list", formula1='"PENDING,VERIFIED,DISCREPANCY_FLAGGED"', allow_blank=True)
    ws.add_data_validation(dv)
    status_col_idx = df.columns.get_loc("verification_status") + 1
    col_letter = get_column_letter(status_col_idx)
    dv.add(f'{col_letter}2:{col_letter}{ws.max_row}')

    wb.save(filepath)
    print(f"Excel file successfully generated at {filepath}")

def parse_excel(filepath="dummy_attendance_records.xlsx"):
    """
    Parses the generated Excel file back into a list of dictionaries matching
    the ngage_client.py payload schema.
    """
    df = pd.read_excel(filepath, sheet_name="Attendance_Summary")
    
    # Handle NaN/Null values if necessary
    df = df.fillna("")
    
    records = df.to_dict(orient="records")
    return records

def update_excel_record(filepath, employee_code, days_worked, leaves, absences):
    """
    Updates an employee's attendance record in the Excel file without losing formatting.
    """
    wb = load_workbook(filepath)
    if "Attendance_Summary" not in wb.sheetnames:
        return False
        
    ws = wb["Attendance_Summary"]
    
    # Find headers to get column indices
    headers = {cell.value: cell.column for cell in ws[1]}
    
    code_col = headers.get("employee_code")
    days_col = headers.get("days_worked")
    leaves_col = headers.get("approved_leaves")
    abs_col = headers.get("unapproved_absences")
    
    if not all([code_col, days_col, leaves_col, abs_col]):
        return False
        
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=code_col).value == employee_code:
            ws.cell(row=row, column=days_col).value = float(days_worked)
            ws.cell(row=row, column=leaves_col).value = float(leaves)
            ws.cell(row=row, column=abs_col).value = float(absences)
            wb.save(filepath)
            return True
            
    return False

if __name__ == "__main__":
    import os
    
    # Generate the excel file
    filepath = "dummy_attendance_records.xlsx"
    generate_excel(filepath)
    
    # Test the parser
    print("-" * 50)
    print("Testing Parser Utility:")
    parsed_data = parse_excel(filepath)
    
    for row in parsed_data[:2]:
        print(row)
